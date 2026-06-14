"""
Excel eksport endpointlari (admin uchun).

GET /export/sales?date_from=&date_to=  → .xlsx fayl
GET /export/points                      → .xlsx fayl
"""
from datetime import datetime, timezone, timedelta
from io import BytesIO

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ..database import get_db
from ..models import User, Point, PointStock, Sale
from ..deps import require_role

router = APIRouter(prefix="/export", tags=["export"])

HEADER_FILL = PatternFill("solid", fgColor="1b8a5a")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_header(ws, headers: list[str], col_widths: list[int]):
    ws.row_dimensions[1].height = 24
    for i, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        ws.column_dimensions[get_column_letter(i)].width = w


def _xlsx_response(wb: openpyxl.Workbook, filename: str) -> StreamingResponse:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _dt(iso: str | None, tz=timezone.utc) -> datetime | None:
    if not iso:
        return None
    return datetime.fromisoformat(iso).replace(tzinfo=tz)


@router.get("/sales")
async def export_sales(
    date_from: str | None = Query(default=None),
    date_to:   str | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
    _admin: User = Depends(require_role("admin")),
):
    """Sotuvlar hisobotini Excel formatida yuklab olish."""
    q = (
        select(Sale, User.full_name, Point.name)
        .join(User, Sale.seller_id == User.id)
        .outerjoin(Point, Sale.point_id == Point.id)
    )
    if date_from:
        df = _dt(date_from)
        if df:
            q = q.where(Sale.created_at >= df)
    if date_to:
        dt = _dt(date_to)
        if dt:
            q = q.where(Sale.created_at < dt + timedelta(days=1))
    q = q.order_by(Sale.created_at.desc())

    rows = (await db.execute(q)).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sotuvlar"
    ws.freeze_panes = "A2"

    headers = ["№", "Sana va vaqt", "Sotuvchi", "Tochka", "Operator", "Tur"]
    widths  = [5,   20,             20,          24,       12,         10]
    _style_header(ws, headers, widths)

    for idx, (sale, seller_name, point_name) in enumerate(rows, start=1):
        dt_local = sale.created_at.strftime("%d.%m.%Y %H:%M") if sale.created_at else ""
        ws.append([
            idx,
            dt_local,
            seller_name or "",
            point_name or "Ofis",
            sale.operator.upper(),
            "Tochka" if sale.source == "point" else "Ofis",
        ])

    # Jami qator
    if rows:
        ws.append([])
        ws.append(["", "", "", "", "Jami:", len(rows)])
        last_row = ws.max_row
        ws.cell(last_row, 5).font = Font(bold=True)
        ws.cell(last_row, 6).font = Font(bold=True)

    today = datetime.now(timezone.utc).strftime("%Y%m%d")
    return _xlsx_response(wb, f"sotuvlar_{today}.xlsx")


@router.get("/points")
async def export_points(
    include_archived: bool = Query(default=False),
    db: AsyncSession = Depends(get_db),
    _admin: User = Depends(require_role("admin")),
):
    """Tochkalar ro'yxatini Excel formatida yuklab olish."""
    q = (
        select(Point, User.full_name)
        .join(User, Point.agent_id == User.id)
    )
    if not include_archived:
        q = q.where(Point.is_archived == False)  # noqa: E712
    q = q.order_by(Point.created_at.desc())

    rows = (await db.execute(q)).all()

    # Tochka stocklarini bir so'rovda olamiz
    point_ids = [p.id for p, _ in rows]
    stock_map: dict[str, dict[str, int]] = {}
    if point_ids:
        sr = await db.execute(
            select(PointStock).where(PointStock.point_id.in_(point_ids))
        )
        for ps in sr.scalars():
            stock_map.setdefault(ps.point_id, {})[ps.operator] = ps.qty

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tochkalar"
    ws.freeze_panes = "A2"

    ops = ["beeline", "ucell", "uzmobile", "mobiuz", "oq"]
    headers = ["№", "Nomi", "Hudud", "Agent", "Ochilgan sana", "Jami SIM", *[o.upper() for o in ops], "Holat"]
    widths  = [5,   22,     20,      18,      16,               10,        *([10] * len(ops)),             10]
    _style_header(ws, headers, widths)

    for idx, (point, agent_name) in enumerate(rows, start=1):
        ps = stock_map.get(point.id, {})
        total = sum(ps.values())
        op_qtys = [ps.get(op, 0) for op in ops]
        ws.append([
            idx,
            point.name,
            point.location,
            agent_name or "",
            point.created_at.strftime("%d.%m.%Y") if point.created_at else "",
            total,
            *op_qtys,
            "Arxiv" if point.is_archived else "Faol",
        ])

    today = datetime.now(timezone.utc).strftime("%Y%m%d")
    return _xlsx_response(wb, f"tochkalar_{today}.xlsx")
